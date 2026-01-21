#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import re
import pandas as pd
from pathlib import Path
import sys

STATUS_ORDER = {"严重": 0, "中等": 1, "轻微": 2, "误报": 3}

def get_plan(status):
    status = status.strip()
    if status == '严重':
        return '立即修改'
    elif status == '中等':
        return '本周'
    elif status == '轻微':
        return '本月'
    elif status == '误报':
        return '不修改'
    return ''

def extract_issues_from_text(html_content):
    issues = []

    # Method 1: Pattern for headers like <hX>🔴 严重: ...</hX>
    html_pattern = re.compile(r'<h[0-9][^>]*>\s*(?:🔴|🟡|🟢)?\s*(严重|中等|轻微|误报)\s*[:：]\s*(.+?)\s*</h[0-9]>')
    matches = html_pattern.findall(html_content)

    # Method 2: Fallback simple pattern
    if not matches:
        simple_pattern = re.compile(r'(?:🔴|🟡|🟢)\s*(严重|中等|轻微|误报)\s*[:：]\s*([^\n<]+)')
        matches = simple_pattern.findall(html_content)

    # Method 3: Broadest fallback
    if not matches:
        broad_pattern = re.compile(r'(严重|中等|轻微|误报)\s*[:：]\s*([^\n<]+)')
        matches = broad_pattern.findall(html_content)

    seen = set()
    for status, problem in matches:
        problem = problem.strip()
        # Clean HTML tags if any remain
        problem = re.sub(r'<[^>]+>', '', problem)

        # Deduplication
        key = f"{status}:{problem}"
        if key not in seen:
            issues.append({
                '问题': problem,
                '问题状态': status,
                '修改计划': get_plan(status)
            })
            seen.add(key)

    return issues

def parse_html_tables(html_path: Path):
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # First try text extraction (works for the header-based format)
        data = extract_issues_from_text(content)

        # If text extraction found nothing, try table parsing (legacy support)
        if not data:
            tables = pd.read_html(str(html_path))
            if tables:
                # Reuse the logic from original script for table parsing if needed
                # For now, let's assume if regex failed, we might look for table structure
                # But typically regex is robust enough for the current report format
                pass

        return data
    except Exception as e:
        print(f"Error parsing HTML: {e}")
        return []

def write_xlsx(rows, out_path: Path):
    if not rows:
        print("No data to write.")
        return

    df = pd.DataFrame(rows)

    # Sort
    # Map status to order value, default to 99 if not found
    df['sort_key'] = df['问题状态'].map(lambda x: STATUS_ORDER.get(x, 99))
    df = df.sort_values('sort_key').drop('sort_key', axis=1)

    # Ensure columns are in correct order
    cols = ['问题', '问题状态', '修改计划']
    # Filter to only these columns if they exist
    df = df[[c for c in cols if c in df.columns]]

    try:
        # Use pandas to write, cleaner dependency management usually
        df.to_excel(out_path, index=False)

        # Optional: formatting with openpyxl if installed, to match original style
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            wb = load_workbook(out_path)
            ws = wb.active

            # Style headers
            for c in ws[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(vertical="center")

            # Set column widths
            widths = [60, 12, 12]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            wb.save(out_path)
        except ImportError:
            pass # openpyxl might not be installed, basic xlsx is fine

    except Exception as e:
        print(f"Error saving Excel file: {e}")
        sys.exit(1)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="输入 HTML 路径")
    ap.add_argument("--output", default="", help="输出 xlsx 路径（默认同名 .xlsx）")
    args = ap.parse_args()

    html_path = Path(args.input)
    if not html_path.exists():
        print(f"Error: File {html_path} not found.")
        sys.exit(1)

    out_path = Path(args.output) if args.output else html_path.with_suffix(".xlsx")

    rows = parse_html_tables(html_path)

    if not rows:
        print("未从 HTML 中提取到有效的问题记录。")
        sys.exit(1)

    write_xlsx(rows, out_path)
    print(f"OK: {html_path} -> {out_path}")

if __name__ == "__main__":
    main()
